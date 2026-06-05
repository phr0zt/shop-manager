from fastapi import APIRouter, Depends, HTTPException, Query, Request, Form
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from typing import List, Optional
from datetime import date, datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.database import get_db
from app.models import Customer, Vehicle, Inspection, InspectionItem, Repair, Job, Payment, JobStatus, RepairStatus
from app.schemas import *

router = APIRouter()


# ---------- Helpers ----------
def get_or_404(db: Session, model, id: int):
    obj = db.query(model).filter(model.id == id).first()
    if not obj:
        raise HTTPException(404, f"{model.__name__} not found")
    return obj


# ---------- Customers ----------
@router.get("/customers", response_class=HTMLResponse)
def customers_page(request: Request, q: str = "", db: Session = Depends(get_db)):
    query = db.query(Customer)
    if q:
        query = query.filter(or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%")))
    customers = query.order_by(Customer.name).all()
    return request.app.templates.TemplateResponse("customers.html", {"request": request, "customers": customers, "q": q})


@router.post("/customers", response_class=HTMLResponse)
def create_customer(request: Request, name: str = Form(...), phone: str = Form(None), email: str = Form(None), address: str = Form(None), notes: str = Form(None), db: Session = Depends(get_db)):
    c = Customer(name=name, phone=phone, email=email, address=address, notes=notes)
    db.add(c)
    db.commit()
    db.refresh(c)
    return HTMLResponse(f'<tr id="customer-{c.id}"><td>{c.name}</td><td>{c.phone or ""}</td><td>{c.email or ""}</td><td><button hx-get="/customers/{c.id}/edit" class="btn">Edit</button> <button hx-delete="/customers/{c.id}" hx-target="#customer-{c.id}" class="btn danger">Delete</button></td></tr>')


@router.get("/customers/{id}/edit", response_class=HTMLResponse)
def edit_customer_form(id: int, request: Request, db: Session = Depends(get_db)):
    c = get_or_404(db, Customer, id)
    return request.app.templates.TemplateResponse("customer_form.html", {"request": request, "customer": c})


@router.put("/customers/{id}", response_class=HTMLResponse)
def update_customer(id: int, request: Request, name: str = Form(...), phone: str = Form(None), email: str = Form(None), address: str = Form(None), notes: str = Form(None), db: Session = Depends(get_db)):
    c = get_or_404(db, Customer, id)
    c.name, c.phone, c.email, c.address, c.notes = name, phone, email, address, notes
    db.commit()
    return HTMLResponse(f'<tr id="customer-{c.id}"><td>{c.name}</td><td>{c.phone or ""}</td><td>{c.email or ""}</td><td><button hx-get="/customers/{c.id}/edit" class="btn">Edit</button> <button hx-delete="/customers/{c.id}" hx-target="#customer-{c.id}" class="btn danger">Delete</button></td></tr>')


@router.delete("/customers/{id}")
def delete_customer(id: int, db: Session = Depends(get_db)):
    c = get_or_404(db, Customer, id)
    db.delete(c)
    db.commit()
    return HTMLResponse("")


# ---------- Vehicles ----------
@router.get("/vehicles", response_class=HTMLResponse)
def vehicles_page(request: Request, customer_id: int = None, db: Session = Depends(get_db)):
    query = db.query(Vehicle)
    if customer_id:
        query = query.filter(Vehicle.customer_id == customer_id)
    vehicles = query.join(Customer).order_by(Customer.name, Vehicle.make, Vehicle.model).all()
    customers = db.query(Customer).order_by(Customer.name).all()
    return request.app.templates.TemplateResponse("vehicles.html", {"request": request, "vehicles": vehicles, "customers": customers, "selected_customer": customer_id})


@router.post("/vehicles", response_class=HTMLResponse)
def create_vehicle(request: Request, customer_id: int = Form(...), make: str = Form(...), model: str = Form(...), year: int = Form(...), vin: str = Form(None), license_plate: str = Form(None), color: str = Form(None), mileage: int = Form(None), notes: str = Form(None), db: Session = Depends(get_db)):
    v = Vehicle(customer_id=customer_id, make=make, model=model, year=year, vin=vin, license_plate=license_plate, color=color, mileage=mileage, notes=notes)
    db.add(v)
    db.commit()
    db.refresh(v)
    customer = db.query(Customer).get(v.customer_id)
    return HTMLResponse(f'<tr id="vehicle-{v.id}"><td>{customer.name}</td><td>{v.year} {v.make} {v.model}</td><td>{v.license_plate or ""}</td><td>{v.mileage or ""}</td><td><a hx-get="/vehicles/{v.id}/inspect" class="btn">Inspect</a> <a hx-get="/vehicles/{v.id}/jobs" class="btn">Jobs</a></td></tr>')


# ---------- Inspections ----------
@router.get("/vehicles/{id}/inspect", response_class=HTMLResponse)
def new_inspection_form(id: int, request: Request, db: Session = Depends(get_db)):
    vehicle = get_or_404(db, Vehicle, id)
    categories = ["Brakes", "Tires", "Fluids", "Suspension", "Steering", "Lights", "Electrical", "Exhaust", "Engine", "Transmission", "Cooling", "Belts/Hoses", "Body/Frame", "Other"]
    return request.app.templates.TemplateResponse("inspection_form.html", {"request": request, "vehicle": vehicle, "categories": categories})


@router.post("/inspections", response_class=HTMLResponse)
def create_inspection(request: Request, vehicle_id: int = Form(...), date: str = Form(...), mileage: int = Form(None), notes: str = Form(None), category: List[str] = Form(...), component: List[str] = Form(...), condition: List[str] = Form(...), notes_item: List[str] = Form(...), recommended_action: List[str] = Form(...), urgency: List[str] = Form(...), estimated_cost: List[float] = Form(...), db: Session = Depends(get_db)):
    insp = Inspection(vehicle_id=vehicle_id, date=datetime.strptime(date, "%Y-%m-%d").date(), mileage=mileage, notes=notes)
    db.add(insp)
    db.flush()
    for i in range(len(category)):
        if component[i].strip():
            item = InspectionItem(
                inspection_id=insp.id, category=category[i], component=component[i],
                condition=condition[i], notes=notes_item[i], recommended_action=recommended_action[i],
                urgency=urgency[i], estimated_cost=estimated_cost[i] or 0
            )
            db.add(item)
    db.commit()
    return HTMLResponse(f'<div class="toast">Inspection saved for vehicle {vehicle_id}</div>')


@router.get("/inspections", response_class=HTMLResponse)
def inspections_list(request: Request, db: Session = Depends(get_db)):
    inspections = db.query(Inspection).join(Vehicle).join(Customer).order_by(Inspection.date.desc()).all()
    return request.app.templates.TemplateResponse("inspections.html", {"request": request, "inspections": inspections})


# ---------- Jobs ----------
@router.get("/jobs", response_class=HTMLResponse)
def jobs_board(request: Request, status: str = None, db: Session = Depends(get_db)):
    query = db.query(Job).join(Customer).join(Vehicle)
    if status:
        query = query.filter(Job.status == status)
    jobs = query.order_by(Job.created_at.desc()).all()
    return request.app.templates.TemplateResponse("jobs.html", {"request": request, "jobs": jobs, "statuses": [s.value for s in JobStatus], "current_status": status})


@router.get("/vehicles/{id}/jobs", response_class=HTMLResponse)
def vehicle_jobs(id: int, request: Request, db: Session = Depends(get_db)):
    vehicle = get_or_404(db, Vehicle, id)
    jobs = db.query(Job).filter(Job.vehicle_id == id).order_by(Job.created_at.desc()).all()
    return request.app.templates.TemplateResponse("vehicle_jobs.html", {"request": request, "vehicle": vehicle, "jobs": jobs})


@router.post("/jobs", response_class=HTMLResponse)
def create_job(request: Request, customer_id: int = Form(...), vehicle_id: int = Form(...), title: str = Form(...), description: str = Form(None), quoted_total: float = Form(0), status: str = Form("quoted"), db: Session = Depends(get_db)):
    job = Job(customer_id=customer_id, vehicle_id=vehicle_id, title=title, description=description, quoted_total=quoted_total, status=status)
    db.add(job)
    db.commit()
    db.refresh(job)
    return HTMLResponse(f'<div class="toast">Job #{job.id} created</div>')


@router.patch("/jobs/{id}/status", response_class=HTMLResponse)
def update_job_status(id: int, status: str = Form(...), db: Session = Depends(get_db)):
    job = get_or_404(db, Job, id)
    job.status = status
    if status == "active" and not job.started_at:
        job.started_at = datetime.now()
    elif status == "completed" and not job.completed_at:
        job.completed_at = datetime.now()
    elif status == "paid" and not job.paid_at:
        job.paid_at = datetime.now()
    db.commit()
    return HTMLResponse(f'<span class="badge {status}">{status}</span>')


# ---------- Repairs ----------
@router.post("/repairs", response_class=HTMLResponse)
def create_repair(request: Request, inspection_item_id: int = Form(...), job_id: int = Form(None), description: str = Form(None), parts_cost: float = Form(0), labor_cost: float = Form(0), labor_hours: float = Form(0), notes: str = Form(None), db: Session = Depends(get_db)):
    repair = Repair(inspection_item_id=inspection_item_id, job_id=job_id, description=description, parts_cost=parts_cost, labor_cost=labor_cost, labor_hours=labor_hours, notes=notes)
    db.add(repair)
    db.commit()
    db.refresh(repair)
    return HTMLResponse(f'<div class="toast">Repair added</div>')


@router.patch("/repairs/{id}/status", response_class=HTMLResponse)
def update_repair_status(id: int, status: str = Form(...), db: Session = Depends(get_db)):
    repair = get_or_404(db, Repair, id)
    repair.status = status
    if status == "done":
        repair.completed_at = datetime.now()
    db.commit()
    return HTMLResponse(f'<span class="badge {status}">{status}</span>')


# ---------- Payments ----------
@router.post("/payments", response_class=HTMLResponse)
def add_payment(request: Request, job_id: int = Form(...), amount: float = Form(...), method: str = Form(None), date: str = Form(...), notes: str = Form(None), db: Session = Depends(get_db)):
    payment = Payment(job_id=job_id, amount=amount, method=method, date=datetime.strptime(date, "%Y-%m-%d").date(), notes=notes)
    db.add(payment)
    # update job actual_total and check if fully paid
    job = db.query(Job).get(job_id)
    total_paid = sum(p.amount for p in job.payments) + amount
    job.actual_total = total_paid
    if total_paid >= job.quoted_total and job.status != "paid":
        job.status = "paid"
        job.paid_at = datetime.now()
    db.commit()
    return HTMLResponse(f'<div class="toast">Payment recorded</div>')


# ---------- Tax Export ----------
@router.get("/export/tax")
def export_tax(year: int = Query(None), db: Session = Depends(get_db)):
    query = db.query(Job).join(Customer).join(Vehicle)
    if year:
        query = query.filter(func.strftime("%Y", Job.created_at) == str(year))
    jobs = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Export"

    headers = ["Date", "Customer", "Vehicle", "Job Title", "Status", "Quoted Total", "Actual Total", "Tax Amount", "Labor Total", "Parts Total", "Paid", "Balance"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, job in enumerate(jobs, 2):
        labor = sum(r.labor_cost for r in job.repairs)
        parts = sum(r.parts_cost for r in job.repairs)
        paid = sum(p.amount for p in job.payments)
        ws.cell(row=row_idx, column=1, value=job.created_at.strftime("%Y-%m-%d"))
        ws.cell(row=row_idx, column=2, value=job.customer.name)
        ws.cell(row=row_idx, column=3, value=f"{job.vehicle.year} {job.vehicle.make} {job.vehicle.model}")
        ws.cell(row=row_idx, column=4, value=job.title)
        ws.cell(row=row_idx, column=5, value=job.status)
        ws.cell(row=row_idx, column=6, value=job.quoted_total)
        ws.cell(row=row_idx, column=7, value=job.actual_total)
        ws.cell(row=row_idx, column=8, value=job.tax_amount)
        ws.cell(row=row_idx, column=9, value=labor)
        ws.cell(row=row_idx, column=10, value=parts)
        ws.cell(row=row_idx, column=11, value=paid)
        ws.cell(row=row_idx, column=12, value=job.quoted_total - paid)

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"tax_export_{year or 'all'}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


# ---------- Dashboard ----------
@router.get("/", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db)):
    stats = {
        "customers": db.query(Customer).count(),
        "vehicles": db.query(Vehicle).count(),
        "jobs_quoted": db.query(Job).filter(Job.status == "quoted").count(),
        "jobs_active": db.query(Job).filter(Job.status == "active").count(),
        "jobs_unpaid": db.query(Job).filter(Job.status.in_(["completed", "unpaid"])).count(),
        "total_revenue": sum(j.actual_total for j in db.query(Job).all()),
    }
    recent_jobs = db.query(Job).join(Customer).join(Vehicle).order_by(Job.created_at.desc()).limit(10).all()
    return request.app.templates.TemplateResponse("dashboard.html", {"request": request, "stats": stats, "recent_jobs": recent_jobs})
